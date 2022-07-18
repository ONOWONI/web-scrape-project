from bs4 import BeautifulSoup
import requests
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


html_text = requests.get('https://www.jumia.com.ng/computing/').text
# wb = Workbook()

# wb = load_workbook('book1.xlsx')

# ws = wb.active

# office = ws['A2':'B10']

# for cell in office:
#     for i in cell:
#         print(f"{i.value}")


content = html_text

soup = BeautifulSoup(content, "lxml")
tags = soup.find_all("div", class_="itm col")
for i in tags:
    name = i.find('div', class_="name")
    if "Hp" in name.text:
        price = i.find('div', class_= "prc")
        link = i.article.a["href"]
        print(f'\n Name : {name.text} \n Price : {price.text}\n link : {link}')